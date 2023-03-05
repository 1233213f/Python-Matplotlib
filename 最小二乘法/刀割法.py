# 拟合曲线
import math
import copy
import random
import numpy as np
import scipy as sp
import matplotlib.pyplot as plt
from scipy.optimize import leastsq
import math
import random
import openpyxl
import numpy as np
import PySimpleGUI as sg
from scipy import integrate
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter, column_index_from_string

def jakd(x, y):

    Yi = np.array(y)
    Xi = np.array(x)

    # 需要拟合的函数func（）指定函数的形状
    def func(p, x):
        a, b = p
        return a*x + b

    # 定义偏差函数，x，y为数组中对应Xi,Yi的值
    def error(p, x, y):
        return func(p, x) - y

    # 设置a，b的初始值，可以任意设定，经过实验，发现p0的值会影响cost的值：Para[1]
    p0 = [5, -8]

    # 把error函数中除了p0以外的参数打包到args中,leastsq()为最小二乘法函数
    Para = leastsq(error, p0, args=(Xi, Yi))
    # 读取结果
    a, b = Para[0]
    return Para[0]

list_y = [18.5, 24.4, 31.0, 2.8, -3]
list_x = [4.690895705, 5.829204678, 7.331880427, 4.698970004, 0.477121255]
a = []
b = []
for i in range(0, 50):
    re = random.randint(0, 4)
    x = copy.deepcopy(list_x)
    y = copy.deepcopy(list_y)
    x.pop(re)
    y.pop(re)
    a.append(jakd(x, y)[0])
    b.append(jakd(x, y)[1])
    # print(jakd(x, y))

a_avg = np.mean(a)
b_avg = np.mean(b)
print('a=', np.mean(a), 'b=', np.mean(b))

# 画样本点
plt.figure(figsize=(8, 6))
plt.scatter(np.array(list_x), np.array(list_y), color='red', label='Sample data', linewidth=2)

# 画拟合直线
x = np.linspace(0, 8, 80)
y = a_avg * x + b_avg

# 绘制拟合曲线
plt.plot(x, y, color='blue', label='Fitting Curve', linewidth=2)
plt.legend()  # 绘制图例

plt.xlabel('D(pc)', fontproperties='simHei', fontsize=12)
plt.ylabel('m-M', fontproperties='simHei', fontsize=12)

plt.show()

# 在内存创建一个工作簿
wb = Workbook()
ws = wb.active
for i in range(0, 50):
    ws.cell(row=i + 1, column=0 + 1).value = a[i]
    ws.cell(row=i + 1, column=1 + 1).value = b[i]
# 保存文档
wb.save('E:\课题&复习资料\大三上\天体物理中的统计方法\作业\y_ab.xlsx')
